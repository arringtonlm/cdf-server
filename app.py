from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import openpyxl
import anthropic
import base64
import io
import os
import json

app = Flask(__name__)

NETLIFY_URL = "https://radiant-daifuku-afd6ca.netlify.app"

CORS(app, resources={r"/*": {"origins": [NETLIFY_URL, "http://localhost:8080", "http://127.0.0.1:8080"]}})

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "CDF_Template.xlsx")


def get_client():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise ValueError("ANTHROPIC_API_KEY environment variable is not set")
    return anthropic.Anthropic(api_key=api_key)


@app.route("/health", methods=["GET"])
def health():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    return jsonify({"status": "ok", "api_key_set": bool(api_key)})


# -- Receipt scanning ---------------------------------------------------------
@app.route("/scan-receipt", methods=["POST", "OPTIONS"])
def scan_receipt():
    if request.method == "OPTIONS":
        return "", 204

    try:
        client = get_client()

        if "image" not in request.files:
            return jsonify({"error": "No image file received"}), 400

        file = request.files["image"]
        image_data = file.read()

        if not image_data:
            return jsonify({"error": "Image file is empty"}), 400

        media_type = file.content_type or "image/jpeg"
        if media_type not in ("image/jpeg", "image/png", "image/gif", "image/webp"):
            media_type = "image/jpeg"

        b64 = base64.standard_b64encode(image_data).decode("utf-8")

        prompt = """You are reading a receipt image to extract expense data.

IMPORTANT CURRENCY RULES:
- If the receipt shows CDF, Fc, FC, or Congolese Francs: currency = "CDF"
- If the receipt shows $, USD, or US Dollars: currency = "USD"
- Do NOT convert between currencies — report the exact amount shown on the receipt
- If both currencies appear (e.g. a bank slip alongside a store receipt), use the currency of the store receipt total

Extract each distinct line item. For receipts showing one total (e.g. a restaurant bill), return ONE item using the total.

Return ONLY a valid JSON array, no markdown, no explanation. Each object:
{
  "description": "short description",
  "date": "DD/MM/YYYY or empty string",
  "qty": 1,
  "unitPrice": <number — exact amount in the receipt currency>,
  "currency": "USD" or "CDF"
}"""

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1000,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
                    {"type": "text", "text": prompt}
                ]
            }]
        )

        text = "".join(block.text for block in message.content if hasattr(block, "text"))
        clean = text.replace("```json", "").replace("```", "").strip()
        items = json.loads(clean)
        return jsonify({"items": items})

    except ValueError as e:
        return jsonify({"error": str(e)}), 503
    except json.JSONDecodeError as e:
        return jsonify({"error": f"Could not parse AI response: {str(e)}"}), 422
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# -- CDF filling --------------------------------------------------------------
@app.route("/fill-cdf", methods=["POST", "OPTIONS"])
def fill_cdf():
    if request.method == "OPTIONS":
        return "", 204

    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No JSON data received"}), 400

        name           = data.get("name", "")
        email          = data.get("email", "")
        phone          = data.get("phone", "")
        location       = data.get("location", "")
        req_num        = data.get("req_num", "")
        date_submitted = data.get("date_submitted", "")
        currency       = data.get("currency", "USD")
        items          = data.get("items", [])

        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb["Cash Disbursement"]

        ws["L1"] = req_num
        ws["L2"] = date_submitted
        ws["L3"] = date_submitted
        ws["C4"] = name
        ws["C5"] = phone
        ws["I5"] = location
        ws["C6"] = email
        ws["H6"] = currency
        ws["I6"] = "X" if currency == "USD" else ""
        ws["J6"] = "X" if currency == "CDF" else ""

        grand_total = 0.0
        for i, item in enumerate(items):
            if i >= 19:
                break
            row        = 22 + i
            desc       = item.get("description", "")
            item_date  = item.get("date", "")
            speedkey   = item.get("speedkey", "")
            qty        = float(item.get("qty", 1))
            unit_price = float(item.get("unitPrice", 0))
            line_total = round(qty * unit_price, 2)
            grand_total += line_total

            if item_date:
                desc = f"{desc} — {item_date}"

            ws[f"B{row}"] = desc
            ws[f"D{row}"] = speedkey
            ws[f"G{row}"] = qty
            ws[f"H{row}"] = unit_price
            ws[f"J{row}"] = line_total
            ws[f"L{row}"] = line_total

        grand_total = round(grand_total, 2)
        ws["J41"] = grand_total
        ws["L41"] = grand_total
        ws["A46"] = name
        ws["C64"] = name
        ws["D52"] = grand_total
        ws["D54"] = grand_total
        ws["D56"] = 0
        ws["D58"] = 0

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_name = name.replace(" ", "_") or "Staff"
        date_str  = date_submitted.replace("/", "").replace("-", "") or "nodate"
        filename  = f"CDF_{safe_name}_{currency}_{date_str}.xlsx"

        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
